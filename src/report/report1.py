
"""
report1 — enriches the ILP Excel artifact with unit descriptions.

Reads:
- outputs/report/optimized_plan_<solve_date>.xlsx   (latest file produced by ilp.py)
- data/raw/df_location_description.csv              (optional – adds unit_description column)

Writes:
- outputs/report/optimized_plan_<solve_date>.xlsx   (same file, Optimization occupancy sheet
                                                      updated with unit_description column)

No CSV is created.
"""

from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


# ------------------ CONFIG ------------------
ROOT = Path(".")
REPORT_DIR = ROOT / "outputs" / "report"


# ------------------ Helpers ------------------
def _find_latest_xlsx(directory: Path) -> Path:
    """Return outputs/report/optimized_plan.xlsx, raising if absent."""
    p = directory / "optimized_plan.xlsx"
    if not p.exists():
        raise FileNotFoundError(
            f"'optimized_plan.xlsx' not found in {directory}. "
            "Run build_and_solve first."
        )
    return p


def generate_report(project_root: Path | None = None) -> Path:
    """Enrich the latest optimized_plan xlsx with unit_description and write it back.

    Parameters
    ----------
    project_root : Path, optional
        Overrides the module-level ROOT so paths resolve correctly when called
        from another working directory (e.g. run_pipeline.py).

    Returns
    -------
    Path to the (updated) xlsx file.
    """
    global ROOT, REPORT_DIR
    if project_root is not None:
        ROOT = Path(project_root)
        REPORT_DIR = ROOT / "outputs" / "report"
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # ------------------ Locate & load the ILP Excel file ------------------
    xlsx_path = _find_latest_xlsx(REPORT_DIR)
    print(f"[report1] Enriching {xlsx_path.name}")

    SHEET = "Optimization occupancy"

    # Read all sheets so we can write them all back (preserve Overflows, Transfers, Heatmap)
    all_sheets: dict[str, pd.DataFrame] = pd.read_excel(
        xlsx_path, sheet_name=None, engine="openpyxl"
    )

    if SHEET not in all_sheets:
        raise ValueError(
            f"Sheet '{SHEET}' not found in {xlsx_path.name}. "
            f"Available sheets: {list(all_sheets.keys())}"
        )

    df = all_sheets[SHEET].copy()
    df.columns = [str(c).strip() for c in df.columns]

    required = {"date", "unit", "beds_used", "beds", "Optimized Occupancy%", "Overflow"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"Sheet '{SHEET}' in {xlsx_path.name} is missing columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    # ------------------ Unit description mapping ------------------
    loc_desc_path = ROOT / "data" / "raw" / "df_location_description.csv"
    if loc_desc_path.exists():
        loc_desc = pd.read_csv(loc_desc_path)
        loc_desc.columns = [c.strip() for c in loc_desc.columns]
        mnemonic_to_name = dict(zip(
            loc_desc["Mnemonic"].astype(str).str.strip(),
            loc_desc["Name"].astype(str).str.strip(),
        ))
        df["unit_description"] = df["unit"].map(mnemonic_to_name).fillna("")
    else:
        print(f"[warn] {loc_desc_path} not found – unit_description will be blank")
        df["unit_description"] = ""

    # Place unit_description right after 'unit'
    cols = list(df.columns)
    if "unit_description" in cols:
        cols.remove("unit_description")
        unit_idx = cols.index("unit")
        cols.insert(unit_idx + 1, "unit_description")
        df = df[cols]

    df = df.sort_values("unit").reset_index(drop=True)
    all_sheets[SHEET] = df

    # ------------------ Write back all sheets (non-Heatmap ones) ------------------
    # Heatmap sheet contains an image; we preserve it by only rewriting data sheets.
    wb = load_workbook(xlsx_path)
    data_sheets = [s for s in all_sheets if s in wb.sheetnames]

    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name in data_sheets:
            all_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    print("[done] Updated:", xlsx_path.resolve())
    return xlsx_path


if __name__ == "__main__":
    generate_report()
