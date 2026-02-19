"""
report3 – generates Heatmap_OptimizedOcc.png

Reads:
- outputs/report/optimized_plan.xlsx   ("Optimization occupancy" sheet)

Writes:
- outputs/report/Heatmap_OptimizedOcc.png
"""

from __future__ import annotations
from pathlib import Path
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


# ------------------ CONFIG ------------------
ROOT = Path(".")
REPORT_DIR = ROOT / "outputs" / "report"


# ------------------ Helpers ------------------
def _find_xlsx(directory: Path) -> Path:
    """Return outputs/report/optimized_plan.xlsx, raising if absent."""
    p = directory / "optimized_plan.xlsx"
    if not p.exists():
        raise FileNotFoundError(
            f"'optimized_plan.xlsx' not found in {directory}. "
            "Run the pipeline first."
        )
    return p


def _write_heatmap_png(
    occ_tab: pd.DataFrame,
    out_png: Path,
    vmin: float = 0.0,
    vmax: float = 110.0,
) -> None:
    day_str = str(occ_tab["date"].iloc[0])
    locations = occ_tab["unit"].astype(str).tolist()

    pct_df = pd.DataFrame({day_str: occ_tab["Optimized Occupancy%"].astype(float).values}, index=locations)
    beds_used_mat = pd.DataFrame({day_str: occ_tab["beds_used"].astype(int).values}, index=locations)
    beds_mat = pd.DataFrame({day_str: occ_tab["beds"].astype(int).values}, index=locations)
    ov_mat = pd.DataFrame({day_str: occ_tab["Overflow"].astype(int).values}, index=locations)

    cmap = mpl.colormaps.get_cmap("RdYlGn_r") if hasattr(mpl, "colormaps") else plt.get_cmap("RdYlGn_r")
    fig, ax = plt.subplots(figsize=(7, max(4, 0.4 * len(pct_df.index) + 1)))
    im = ax.imshow(pct_df.values.astype(float), aspect="auto", cmap=cmap, norm=Normalize(vmin=vmin, vmax=vmax))

    ax.set_yticks(range(len(pct_df.index)))
    ax.set_yticklabels(pct_df.index)
    ax.set_xticks(range(len(pct_df.columns)))
    ax.set_xticklabels(pct_df.columns, rotation=0)

    for i in range(len(pct_df.index)):
        for j in range(len(pct_df.columns)):
            if pd.isna(pct_df.iloc[i, j]):
                continue
            valp = float(pct_df.iloc[i, j])
            bu_i = int(beds_used_mat.iloc[i, j])
            bd_i = int(beds_mat.iloc[i, j])
            ov_i = int(ov_mat.iloc[i, j])
            label = f"{valp:.0f}%\n{bu_i}/{bd_i}" + (f" +{ov_i}" if ov_i > 0 else "")
            ax.text(j, i, label, ha="center", va="center", fontsize=8,
                    color=("black" if valp < 80 else "white"))

    ax.set_title(f"Optimized occupancy (beds_used/beds + overflow) — {day_str}")
    cbar = plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cbar.set_label("Optimized occupancy %")
    plt.tight_layout()

    out_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(out_png, dpi=160)
    plt.close(fig)
    print(f"[report3] Wrote PNG -> {out_png.resolve()}")


def generate_heatmap_png(project_root: Path | None = None) -> Path:
    """Read the latest optimized_plan xlsx and render a standalone PNG heatmap.

    Parameters
    ----------
    project_root : Path, optional
        Overrides the module-level ROOT so paths resolve correctly when called
        from another working directory (e.g. run_pipeline.py).

    Returns
    -------
    Path to the generated PNG file.
    """
    global ROOT, REPORT_DIR
    if project_root is not None:
        ROOT = Path(project_root)
        REPORT_DIR = ROOT / "outputs" / "report"
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_path = _find_xlsx(REPORT_DIR)
    print(f"[report3] Reading {xlsx_path.name}")

    df = pd.read_excel(xlsx_path, sheet_name="Optimization occupancy", engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"date", "unit", "beds_used", "beds", "Optimized Occupancy%", "Overflow"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"Sheet 'Optimization occupancy' in {xlsx_path.name} is missing columns: {missing}. "
            f"Found: {list(df.columns)}"
        )

    df["beds_used"] = pd.to_numeric(df["beds_used"], errors="coerce").fillna(0).astype(int)
    df["beds"] = pd.to_numeric(df["beds"], errors="coerce").fillna(0).astype(int)
    df["Optimized Occupancy%"] = pd.to_numeric(df["Optimized Occupancy%"], errors="coerce").fillna(0.0)
    df["Overflow"] = pd.to_numeric(df["Overflow"], errors="coerce").fillna(0).astype(int)

    out_png = REPORT_DIR / "Heatmap_OptimizedOcc.png"

    _write_heatmap_png(df, out_png)
    _embed_png_in_xlsx(xlsx_path, out_png)
    return out_png


def _embed_png_in_xlsx(xlsx_path: Path, png_path: Path) -> None:
    """Insert *png_path* into the Heatmap sheet of *xlsx_path* using openpyxl."""
    wb = load_workbook(xlsx_path)

    if "Heatmap" in wb.sheetnames:
        ws = wb["Heatmap"]
        # Remove any existing images so we don't accumulate duplicates
        ws._images = []
    else:
        ws = wb.create_sheet("Heatmap")
        ws["A1"] = f"Optimized occupancy heatmap"
        ws["A2"] = "Cell label shows: % and occ/beds (+overflow if any)"

    img = XLImage(str(png_path))
    img.anchor = "A4"
    ws.add_image(img)

    wb.save(xlsx_path)
    print(f"[report3] Embedded PNG into {xlsx_path.name} (Heatmap sheet)")


if __name__ == "__main__":
    generate_heatmap_png()
